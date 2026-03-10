"""
╔══════════════════════════════════════════════════════════════════╗
║       GOOGLE MAPS SCRAPER v3 — Selenium Edition                 ║
║                                                                  ║
║  INSTALAÇÃO (apenas 1 comando):                                  ║
║  pip install selenium webdriver-manager openpyxl requests        ║
║              beautifulsoup4 lxml                                 ║
║                                                                  ║
║  REQUISITO: Google Chrome instalado no PC (gratuito)            ║
╚══════════════════════════════════════════════════════════════════╝
"""
#oi sou o vininicus amigo do fefe q fez o codigo, ele mandoumuito bem
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import time
import re
import unicodedata
import sqlite3
import random
import queue
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

# ── Verificação de dependências ──────────────────────────────────────────────
MISSING = []
try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.options import Options
    from webdriver_manager.chrome import ChromeDriverManager
    from selenium.webdriver.chrome.service import Service
except ImportError:
    MISSING.append("selenium webdriver-manager")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    MISSING.append("openpyxl")

try:
    import requests
    from bs4 import BeautifulSoup
except ImportError:
    MISSING.append("requests beautifulsoup4 lxml")

# cores
COR_BG       = "#f0f0f0"
COR_CARD     = "#e4e4e4"
COR_BORDA    = "#aaaaaa"
COR_ACCENT   = "#336699"
COR_ACCENT2  = "#2e7d32"
COR_DANGER   = "#cc0000"
COR_TEXTO    = "#111111"
COR_SUBTEXTO = "#555555"
COR_INPUT    = "#ffffff"
COR_LOG_BG   = "#1a1a1a"



# ════════════════════════════════════════════════════════════════════════════
#  BANCO DE HISTÓRICO — evita repetir resultados entre sessões
# ════════════════════════════════════════════════════════════════════════════
class HistoricoDB:
    """
    Banco SQLite local que guarda a URL de todo estabelecimento já aprovado.
    Fica salvo em scraper_historico.db na mesma pasta do programa.
    """
    DB_FILE = Path(__file__).parent / "scraper_historico.db"

    def __init__(self):
        self.conn = sqlite3.connect(str(self.DB_FILE), check_same_thread=False)
        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS aprovados (
                url      TEXT PRIMARY KEY,
                nome     TEXT,
                keyword  TEXT,
                regiao   TEXT,
                data     TEXT
            )
        """)
        self.conn.commit()

    def ja_existe(self, url: str) -> bool:
        cur = self.conn.execute("SELECT 1 FROM aprovados WHERE url = ?", (url,))
        return cur.fetchone() is not None

    def registrar(self, url: str, nome: str, keyword: str, regiao: str):
        try:
            self.conn.execute(
                "INSERT OR IGNORE INTO aprovados (url, nome, keyword, regiao, data) VALUES (?,?,?,?,?)",
                (url, nome, keyword, regiao, datetime.now().strftime("%Y-%m-%d %H:%M"))
            )
            self.conn.commit()
        except Exception:
            pass

    def total(self) -> int:
        cur = self.conn.execute("SELECT COUNT(*) FROM aprovados")
        return cur.fetchone()[0]

    def limpar(self):
        self.conn.execute("DELETE FROM aprovados")
        self.conn.commit()

    def close(self):
        try:
            self.conn.close()
        except Exception:
            pass


# ════════════════════════════════════════════════════════════════════════════
#  SCRAPER  — v4: JavaScript DOM + regex no texto, sem CSS frágil
# ════════════════════════════════════════════════════════════════════════════
class MapsScraper:
    def __init__(self, log_cb=None, progress_cb=None):
        self.log       = log_cb      or print
        self.progress  = progress_cb or (lambda v, t: None)
        self.stop_flag = False
        self.results   = []
        self.driver    = None
        self.db        = HistoricoDB()

    def stop(self):
        self.stop_flag = True

    # ── Normaliza texto para comparação ──────────────────────────────────────
    @staticmethod
    def _norm(txt):
        import unicodedata
        txt = txt.lower()
        txt = unicodedata.normalize("NFD", txt)
        return "".join(c for c in txt if unicodedata.category(c) != "Mn")

    # ── E-mail via site ───────────────────────────────────────────────────────
    def _email_do_site(self, url):
        if not url or not url.startswith("http"):
            return ""
        try:
            r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=8)
            soup = BeautifulSoup(r.text, "lxml")
            # Prioriza mailto: links
            for a in soup.find_all("a", href=True):
                if a["href"].startswith("mailto:"):
                    e = a["href"][7:].split("?")[0].strip()
                    if "@" in e and "." in e.split("@")[1]:
                        return e
            # Fallback: regex no texto
            for e in re.findall(r"[\w.+\-]+@[\w.\-]+\.[a-zA-Z]{2,}", soup.get_text(" ")):
                dom = e.split("@")[1].lower()
                if not any(x in dom for x in ["example","wix","wordpress","seusite","domain"]):
                    return e
        except Exception:
            pass
        return ""

    # ── WhatsApp ──────────────────────────────────────────────────────────────
    def _whatsapp(self, tel):
        """
        Gera link wa.me/ usando o número exatamente como veio do Maps.
        Não assume nem adiciona código de país.
        """
        if not tel:
            return ""
        nums = re.sub(r"\D", "", tel)
        if len(nums) >= 7:
            return f"https://wa.me/{nums}"
        return ""

    # ── Geocodificação ────────────────────────────────────────────────────────
    def _geocodificar(self, regiao):
        try:
            r = requests.get(
                "https://nominatim.openstreetmap.org/search",
                params={"q": regiao, "format": "json", "limit": 1},
                headers={"User-Agent": "MapsScraper/4.0"},
                timeout=8,
            )
            data = r.json()
            if data:
                lat = float(data[0]["lat"])
                lon = float(data[0]["lon"])
                self.log(f"   📍 {data[0].get('display_name','')[:70]}", "sub")
                return lat, lon
        except Exception as e:
            self.log(f"   ⚠ Geocodificação falhou: {e}", "warn")
        return None

    # ── Monta URL do Maps centrada na região ─────────────────────────────────
    def _url_busca(self, keyword, regiao, coords, zoom=13):
        kw = requests.utils.quote(keyword)
        if coords:
            lat, lon = coords
            return f"https://www.google.com/maps/search/{kw}/@{lat},{lon},{zoom}z"
        q = requests.utils.quote(f"{keyword}, {regiao}")
        return f"https://www.google.com/maps/search/{q}"

    # ── Verifica se endereço pertence à região ────────────────────────────────
    # Países/continentes: endereços locais nunca contêm esses termos
    REGIOES_AMPLAS = {
        "brasil", "brazil", "eua", "usa", "estados unidos", "united states",
        "canada", "australia", "franca", "france", "espanha", "spain",
        "portugal", "argentina", "chile", "colombia", "mexico", "peru",
        "alemanha", "germany", "italia", "italy", "japao", "japan",
        "china", "india", "russia", "africa", "europa", "europe",
        "america", "asia", "oceania",
    }

    # Estados BR: nome normalizado → sigla que aparece nos endereços
    ESTADOS_BR = {
        "acre": "ac", "alagoas": "al", "amapa": "ap", "amazonas": "am",
        "bahia": "ba", "ceara": "ce", "distrito federal": "df",
        "espirito santo": "es", "goias": "go", "maranhao": "ma",
        "mato grosso do sul": "ms", "mato grosso": "mt", "minas gerais": "mg",
        "para": "pa", "paraiba": "pb", "parana": "pr", "pernambuco": "pe",
        "piaui": "pi", "rio de janeiro": "rj", "rio grande do norte": "rn",
        "rio grande do sul": "rs", "rondonia": "ro", "roraima": "rr",
        "santa catarina": "sc", "sao paulo": "sp", "sergipe": "se",
        "tocantins": "to",
    }

    def _na_regiao(self, endereco, regiao):
        """
        Verifica se o endereço pertence à região pedida.
        - País/continente → não filtra (endereços locais não contêm "Brasil" etc.)
        - Estado BR (ex: "Ceará") → aceita tanto "ceara" quanto "- ce" no endereço
        - Cidade/bairro → verifica presença no endereço
        """
        if not endereco:
            return True

        regiao_n = self._norm(regiao.strip())
        partes = [p.strip() for p in regiao_n.split(",") if p.strip()]

        # Se todas as partes são países/continentes → aprova direto
        todas_amplas = all(
            any(a in parte for a in self.REGIOES_AMPLAS)
            for parte in partes
        )
        if todas_amplas:
            return True

        end_n = self._norm(endereco)

        for parte in partes:
            # Ignora partes que são países/continentes
            if any(a in parte for a in self.REGIOES_AMPLAS):
                continue

            # Verifica se é nome de estado BR — aceita também a sigla
            sigla = self.ESTADOS_BR.get(parte)
            if sigla:
                # Endereço contém o nome do estado OU a sigla (ex: "- ce" ou "ce,")
                if parte in end_n or re.search(r'\b' + sigla + r'\b', end_n):
                    continue
                else:
                    return False

            # Região composta (ex: "rio de janeiro") — testa nome e sigla
            for nome_estado, sig in self.ESTADOS_BR.items():
                if nome_estado in parte:
                    restante = parte.replace(nome_estado, "").strip()
                    estado_ok = nome_estado in end_n or re.search(r'\b' + sig + r'\b', end_n)
                    if not estado_ok:
                        return False
                    if restante:
                        palavras = [w for w in restante.split() if len(w) > 2]
                        if palavras and not all(w in end_n for w in palavras):
                            return False
                    break
            else:
                # Termo genérico (cidade, bairro, país estrangeiro)
                palavras = [w for w in parte.split() if len(w) > 2]
                if palavras and not all(w in end_n for w in palavras):
                    return False

        return True

    # ── Chrome ────────────────────────────────────────────────────────────────
    # User-agents reais de browsers — rotaciona por instância
    USER_AGENTS = [
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Windows NT 11.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    ]

    def _init_driver(self, worker_id=0, headless=False):
        """
        Cria instância Chrome com anti-detecção:
        - User-agent rotativo por worker
        - Remove todos os sinais de automação
        - Perfil isolado por worker (cookies/sessão separados)
        - Headless opcional (mais rápido, menor risco de bloqueio visual)
        """
        opts = Options()
        # Perfil isolado por worker — cada Chrome parece um usuário diferente
        profile_dir = Path(__file__).parent / f"chrome_profile_w{worker_id}"
        profile_dir.mkdir(exist_ok=True)
        opts.add_argument(f"--user-data-dir={profile_dir}")

        # Anti-detecção
        opts.add_argument("--disable-blink-features=AutomationControlled")
        opts.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
        opts.add_experimental_option("useAutomationExtension", False)
        opts.add_argument(f"--user-agent={self.USER_AGENTS[worker_id % len(self.USER_AGENTS)]}")
        opts.add_argument("--disable-notifications")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-dev-shm-usage")
        opts.add_argument("--disable-gpu")
        opts.add_argument("--lang=pt-BR,pt;q=0.9")
        opts.add_argument("--window-size=1280,900")

        if headless:
            opts.add_argument("--headless=new")

        driver = webdriver.Chrome(
            service=Service(ChromeDriverManager().install()), options=opts
        )

        # Remove propriedades de automação via JS
        driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {"source": """
            Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
            Object.defineProperty(navigator, 'plugins', {get: () => [1,2,3,4,5]});
            Object.defineProperty(navigator, 'languages', {get: () => ['pt-BR','pt','en-US','en']});
            window.chrome = {runtime: {}};
        """})

        return driver

    def _pausa_humana(self, minimo=0.8, maximo=2.2):
        """Delay aleatório que imita comportamento humano."""
        time.sleep(random.uniform(minimo, maximo))

    # ════════════════════════════════════════════════════════════════════════
    #  COLETA DE LINKS — via JavaScript puro, não depende de seletores CSS
    # ════════════════════════════════════════════════════════════════════════
    def _coletar_links(self):
        """
        Usa JavaScript para pegar TODOS os hrefs que contêm /maps/place/
        diretamente do DOM. Não depende de classes CSS que mudam.
        Retorna set de URLs limpas.
        """
        try:
            hrefs = self.driver.execute_script("""
                var links = document.querySelectorAll('a[href*="/maps/place/"]');
                var result = [];
                links.forEach(function(a) {
                    var h = a.href || "";
                    // Filtra só links de estabelecimento (não fotos, não reviews)
                    if (h.includes("/maps/place/") && !h.includes("/photos/") && !h.includes("/reviews/")) {
                        // Limpa parâmetros de rastreamento mas mantém o link base
                        var clean = h.split("?")[0];
                        result.push(clean);
                    }
                });
                // Remove duplicatas
                return [...new Set(result)];
            """)
            return set(hrefs or [])
        except Exception:
            return set()

    def _scroll_lista(self):
        """
        Scroll no painel esquerdo de resultados do Maps.
        Tenta o painel lateral primeiro, fallback para scroll da janela.
        """
        try:
            # O painel de resultados do Maps tem role="main" ou é a div scrollável à esquerda
            scrolled = self.driver.execute_script("""
                // Tenta encontrar o painel de resultados scrollável
                var seletores = [
                    'div[role="feed"]',
                    'div[aria-label*="resultado"]',
                    'div[aria-label*="result"]',
                    'div.m6QErb[aria-label]',
                    'div.m6QErb',
                ];
                for (var i = 0; i < seletores.length; i++) {
                    var el = document.querySelector(seletores[i]);
                    if (el && el.scrollHeight > el.clientHeight) {
                        el.scrollBy(0, 3000);
                        return "feed:" + seletores[i];
                    }
                }
                // Fallback: scroll na página
                window.scrollBy(0, 3000);
                return "window";
            """)
            return scrolled
        except Exception:
            return "err"

    def _fim_de_lista(self):
        """Detecta se o Maps exibiu mensagem de fim de resultados."""
        try:
            body = self.driver.find_element(By.TAG_NAME, "body").text.lower()
            fim_msgs = [
                "chegou ao final", "you've reached the end",
                "fim da lista", "no more results",
                "nenhum resultado", "no results found",
                "não foram encontrados", "we couldn't find"
            ]
            return any(m in body for m in fim_msgs)
        except Exception:
            return False

    def _aguardar_resultados(self, timeout=12):
        """Aguarda aparecer pelo menos 1 link de estabelecimento na página."""
        inicio = time.time()
        while time.time() - inicio < timeout:
            if self._coletar_links():
                return True
            time.sleep(0.8)
        return False

    # ════════════════════════════════════════════════════════════════════════
    #  EXTRAÇÃO DE DADOS DA FICHA
    #  Estratégia: texto da página + regex (robusto a mudanças de HTML)
    # ════════════════════════════════════════════════════════════════════════
    def _extrair_ficha(self):
        d = self.driver

        # Aguarda o nome aparecer (H1)
        try:
            WebDriverWait(d, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "h1"))
            )
        except Exception:
            pass
        time.sleep(1.5)

        # ── Pega o texto completo da ficha ───────────────────────────────────
        try:
            page_text = d.find_element(By.TAG_NAME, "body").text
        except Exception:
            page_text = ""

        # ── Nome (H1 — o mais confiável) ─────────────────────────────────────
        nome = ""
        try:
            for sel in ["h1", "h1.DUwDvf", "h1.fontHeadlineLarge"]:
                els = d.find_elements(By.CSS_SELECTOR, sel)
                for el in els:
                    t = el.text.strip()
                    if t and len(t) > 1:
                        nome = t
                        break
                if nome:
                    break
        except Exception:
            pass

        # ── Categoria ────────────────────────────────────────────────────────
        # O Maps exibe a categoria logo abaixo do nome, em botão ou span pequeno
        categoria = ""
        try:
            for sel in ["button.DkEaL", "button[jsaction*='category']",
                        "span.mgr77e", ".fontBodyMedium button"]:
                els = d.find_elements(By.CSS_SELECTOR, sel)
                for el in els:
                    t = el.text.strip()
                    if t and len(t) < 80:
                        categoria = t
                        break
                if categoria:
                    break
        except Exception:
            pass

        # ── Estrelas — aria-label é o mais confiável ─────────────────────────
        stars = 0.0
        try:
            # Busca qualquer span/div com aria-label contendo "estrela" ou "star"
            for sel in [
                '[aria-label*="estrela"]',
                '[aria-label*="star"]',
                '[aria-label*="rating"]',
                '[aria-label*="nota"]',
            ]:
                els = d.find_elements(By.CSS_SELECTOR, sel)
                for el in els:
                    lbl = el.get_attribute("aria-label") or ""
                    m = re.search(r'(\d[,.]\d|[1-5])', lbl)
                    if m:
                        v = float(m.group(1).replace(",", "."))
                        if 1.0 <= v <= 5.0:
                            stars = v
                            break
                if stars:
                    break
        except Exception:
            pass

        # Fallback: regex no texto da página — procura padrão "4,5 (1.234)"
        if not stars:
            m = re.search(r'\b([1-5][,.]\d)\s*\(', page_text)
            if m:
                try:
                    stars = float(m.group(1).replace(",", "."))
                except Exception:
                    pass

        # ── Avaliações ───────────────────────────────────────────────────────
        reviews = 0
        try:
            for sel in [
                '[aria-label*="avalia"]',
                '[aria-label*="review"]',
                '[aria-label*="opini"]',
            ]:
                els = d.find_elements(By.CSS_SELECTOR, sel)
                for el in els:
                    lbl = el.get_attribute("aria-label") or ""
                    # "1.234 avaliações" ou "1,234 reviews"
                    m = re.search(r'([\d.,]+)\s*(?:avalia|review|opini)', lbl, re.I)
                    if m:
                        v = int(re.sub(r"\D", "", m.group(1)))
                        if v > 0:
                            reviews = v
                            break
                if reviews:
                    break
        except Exception:
            pass

        # Fallback: regex no texto — "(1.234 avaliações)" ou "(1,234)"
        if not reviews:
            # Padrão "(número)" próximo à nota
            m = re.search(r'\b[1-5][,.]\d\s*\(([\d.,]+)\)', page_text)
            if not m:
                m = re.search(r'([\d.]+)\s+avalia', page_text, re.I)
            if m:
                try:
                    reviews = int(re.sub(r"\D", "", m.group(1)))
                except Exception:
                    pass

        # ── Endereço ─────────────────────────────────────────────────────────
        endereco = ""
        try:
            # data-item-id="address" é o mais confiável historicamente
            el = d.find_element(By.CSS_SELECTOR, '[data-item-id="address"]')
            endereco = (el.get_attribute("aria-label") or el.text).strip()
            endereco = re.sub(r"(?i)^endere[çc]o[:\s]*", "", endereco).strip()
        except Exception:
            pass
        if not endereco:
            # Fallback: regex no texto da página — procura padrão de endereço BR
            m = re.search(
                r'(?:Rua|Av(?:enida)?|Al(?:ameda)?|R\.|Estrada|Rod(?:ovia)?|Praça)[^\n]{5,80}',
                page_text, re.I
            )
            if m:
                endereco = m.group(0).strip()

        # ── Telefone ─────────────────────────────────────────────────────────
        telefone = ""
        try:
            el = d.find_element(By.CSS_SELECTOR, '[data-item-id^="phone"]')
            telefone = (el.get_attribute("aria-label") or el.text).strip()
            telefone = re.sub(r"(?i)^telefone[:\s]*", "", telefone).strip()
        except Exception:
            pass
        if not telefone:
            # Regex universal de telefone (BR e internacional)
            m = re.search(r'(?:\+\d{1,3}[\s\-]?)?\(?\d{2,3}\)?[\s\-]?\d{4,5}[\s\-]\d{4}', page_text)
            if m:
                telefone = m.group(0).strip()

        # ── Site ─────────────────────────────────────────────────────────────
        site = ""
        try:
            el = d.find_element(By.CSS_SELECTOR, '[data-item-id="authority"]')
            site = el.get_attribute("href") or ""
        except Exception:
            pass
        if not site:
            try:
                # Qualquer link externo na ficha que não seja do google
                els = d.find_elements(By.CSS_SELECTOR, "a[href^='http']")
                for el in els:
                    href = el.get_attribute("href") or ""
                    if href.startswith("http") and "google" not in href and "maps" not in href:
                        site = href
                        break
            except Exception:
                pass

        return dict(
            nome=nome, categoria=categoria, endereco=endereco,
            telefone=telefone, site=site, stars=stars, reviews=reviews
        )

    # ════════════════════════════════════════════════════════════════════════
    #  SCROLL WORKER — W0 de cada keyword: só faz scroll e alimenta a fila
    # ════════════════════════════════════════════════════════════════════════
    def _scroll_worker(self, worker_id, keyword, regiao, coords,
                       link_queue, scroll_done, results_lock,
                       global_vistos, headless):
        """
        Abre o Maps, faz scroll infinito e coloca links novos na fila.
        Sinaliza scroll_done quando a lista acabar ou stop_flag.
        """
        w = MapsScraper(log_cb=self.log, progress_cb=None)
        try:
            self.log(f"[W{worker_id}] 🖱 scroll worker iniciando...", "sub")
            w.driver = w._init_driver(worker_id=worker_id, headless=headless)
            w._pausa_humana(1.0 + worker_id * 0.3, 2.0)

            zooms = [14, 12, 10, 8] if ("," in regiao or len(regiao.split()) >= 3) else [10, 8, 6]

            for zoom in zooms:
                if self.stop_flag:
                    break

                url = w._url_busca(keyword, regiao, coords, zoom=zoom)
                zoom_label = {14:"cidade",12:"região",10:"estado",8:"país",6:"continental"}.get(zoom, str(zoom))
                self.log(f"[W{worker_id}] 🗺 {keyword!r} zoom:{zoom_label}", "info")

                w.driver.get(url)
                w._pausa_humana(2.5, 4.0)

                if not w._aguardar_resultados(timeout=12):
                    self.log(f"[W{worker_id}] ⚠ sem resultados (zoom {zoom_label})", "warn")
                    continue

                vistos_local = set()
                sem_novos = 0

                while not self.stop_flag and sem_novos < 10:
                    novos = w._coletar_links() - vistos_local
                    with results_lock:
                        novos -= global_vistos

                    if novos:
                        sem_novos = 0
                        vistos_local.update(novos)
                        for link in novos:
                            link_queue.put(link)
                        self.log(f"[W{worker_id}] 📥 +{len(novos)} links na fila (total fila: {link_queue.qsize()})", "sub")
                    else:
                        sem_novos += 1

                    if w._fim_de_lista():
                        self.log(f"[W{worker_id}] 📌 fim da lista (zoom {zoom_label})", "sub")
                        break

                    w._scroll_lista()
                    w._pausa_humana(1.5, 2.5)

        except Exception as e:
            self.log(f"[W{worker_id}] ❌ scroll worker erro: {e}", "erro")
        finally:
            try: w.driver.quit()
            except Exception: pass
            scroll_done.set()
            self.log(f"[W{worker_id}] scroll encerrado", "sub")

    # ════════════════════════════════════════════════════════════════════════
    #  FICHA WORKER — consome links da fila e extrai dados das fichas
    # ════════════════════════════════════════════════════════════════════════
    def _ficha_worker(self, worker_id, regiao, min_stars, min_reviews,
                      meta_por_kw, keyword, link_queue, scroll_done,
                      results_lock, global_vistos, global_aprovados_counter,
                      meta_total, aprovados_kw_counter, headless):
        """
        Consome links da fila, abre cada ficha, filtra e aprova.
        Para quando meta atingida ou fila vazia + scroll encerrado.
        """
        w = MapsScraper(log_cb=self.log, progress_cb=None)
        resultados_locais = []

        try:
            self.log(f"[W{worker_id}] 📋 ficha worker iniciando...", "sub")
            w.driver = w._init_driver(worker_id=worker_id, headless=headless)
            w._pausa_humana(1.5 + worker_id * 0.4, 2.5)

            while not self.stop_flag:
                # Checa se atingiu a meta
                with results_lock:
                    if aprovados_kw_counter[0] >= meta_por_kw:
                        break

                # Pega próximo link da fila
                try:
                    link = link_queue.get(timeout=3)
                except Exception:
                    # Fila vazia — verifica se scroll terminou
                    if scroll_done.is_set() and link_queue.empty():
                        break
                    continue

                # Deduplicação
                with results_lock:
                    if link in global_vistos or self.db.ja_existe(link):
                        self.log(f"[W{worker_id}] 🔁 já visto", "sub")
                        link_queue.task_done()
                        continue

                try:
                    w.driver.get(link)
                    w._pausa_humana(1.2, 2.2)
                    dados = w._extrair_ficha()

                    nome    = dados["nome"] or "(sem nome)"
                    stars   = dados["stars"]
                    reviews = dados["reviews"]
                    end     = dados["endereco"]

                    if not w._na_regiao(end, regiao):
                        self.log(f"[W{worker_id}] 🚫 {nome} — fora da região", "warn")
                        w.driver.back(); w._pausa_humana(0.6, 1.2)
                        link_queue.task_done()
                        continue
                    if stars > 0 and stars < min_stars:
                        self.log(f"[W{worker_id}] ⏭ {nome} | {stars:.1f}⭐", "sub")
                        w.driver.back(); w._pausa_humana(0.6, 1.2)
                        link_queue.task_done()
                        continue
                    if reviews > 0 and reviews < min_reviews:
                        self.log(f"[W{worker_id}] ⏭ {nome} | {reviews}aval.", "sub")
                        w.driver.back(); w._pausa_humana(0.6, 1.2)
                        link_queue.task_done()
                        continue

                    # ── APROVADO ─────────────────────────────────────────────
                    with results_lock:
                        if aprovados_kw_counter[0] >= meta_por_kw:
                            link_queue.task_done()
                            break
                        global_vistos.add(link)
                        self.db.registrar(link, nome, keyword, regiao)
                        aprovados_kw_counter[0] += 1
                        global_aprovados_counter[0] += 1
                        tot = global_aprovados_counter[0]
                        kw_tot = aprovados_kw_counter[0]
                        self.progress(
                            min(tot / meta_total * 100, 99),
                            f"{tot} aprovados  [W{worker_id}: {kw_tot}/{meta_por_kw}]"
                        )

                    email = self._email_do_site(dados["site"]) if dados["site"] else ""

                    resultados_locais.append({
                        "Nome":       nome,
                        "E-mail":     email,
                        "Telefone":   dados["telefone"],
                        "WhatsApp":   self._whatsapp(dados["telefone"]),
                        "Categoria":  dados["categoria"],
                        "Endereço":   end,
                        "Site":       dados["site"],
                        "Estrelas":   stars,
                        "Avaliações": reviews,
                        "Keyword":    keyword,
                        "URL Maps":   link,
                    })

                    st = f"{stars:.1f}⭐" if stars > 0 else "s/nota"
                    rv = f"{reviews:,}aval.".replace(",", ".") if reviews > 0 else "s/aval."
                    self.log(
                        f"[W{worker_id}] ✅ [{kw_tot}/{meta_por_kw}] {nome} | {st} | {rv}"
                        + (" | 📧" if email else ""), "ok"
                    )
                    w.driver.back()
                    w._pausa_humana(1.2, 2.5)

                except Exception as e:
                    self.log(f"[W{worker_id}] ⚠ {e}", "warn")
                    try: w.driver.back(); w._pausa_humana(0.5, 1.0)
                    except Exception: pass

                link_queue.task_done()

        except Exception as e:
            self.log(f"[W{worker_id}] ❌ ficha worker erro: {e}", "erro")
        finally:
            try: w.driver.quit()
            except Exception: pass

        return resultados_locais

    # ════════════════════════════════════════════════════════════════════════
    #  SCRAPE PRINCIPAL
    # ════════════════════════════════════════════════════════════════════════
    def scrape(self, keywords, regiao, min_stars, min_reviews, meta_por_kw,
               save_path, num_workers=1, headless=False):
        self.stop_flag = False
        self.results   = []

        if MISSING:
            self.log(f"❌ faltam dependências: pip install {' '.join(MISSING)}", "erro")
            return []

        self.log(f'🌎 Geocodificando: "{regiao}"...', "info")
        coords = self._geocodificar(regiao)
        if coords:
            self.log(f"   ✅ lat={coords[0]:.4f} lon={coords[1]:.4f}", "ok")
        else:
            self.log("   ⚠ sem coordenadas — fallback textual", "warn")

        total_kw   = len(keywords)
        meta_total = total_kw * meta_por_kw

        # num_workers = total de Chromes por keyword
        # 1 faz scroll + (num_workers-1) processam fichas
        # Se num_workers=1: 1 faz scroll, 1 ficha (mesmo Chrome não funciona bem,
        # então usamos 1 scroll + 1 ficha = 2 Chromes quando workers>=1)
        ficha_workers = max(1, num_workers - 1)  # pelo menos 1 ficha worker
        total_chromes_por_kw = 1 + ficha_workers  # 1 scroll + N ficha

        self.log(
            f"⚡ {num_workers} worker(s) | {total_chromes_por_kw} Chrome(s)/keyword "
            f"| {total_kw} keyword(s) | meta: {meta_total}", "system"
        )
        if headless:
            self.log("   👻 headless ativo", "sub")

        results_lock             = threading.Lock()
        global_vistos            = set()
        global_aprovados_counter = [0]

        try:
            for ki, keyword in enumerate(keywords):
                if self.stop_flag:
                    break

                self.log(f"\n── keyword [{ki+1}/{total_kw}]: {keyword!r} ──", "system")

                link_queue          = queue.Queue()
                scroll_done         = threading.Event()
                aprovados_kw_counter = [0]

                # IDs únicos de worker para perfis Chrome separados
                # offset para não reusar perfis entre keywords
                id_offset = ki * total_chromes_por_kw

                with ThreadPoolExecutor(max_workers=total_chromes_por_kw) as ex:
                    # 1 scroll worker (W0 da keyword)
                    scroll_future = ex.submit(
                        self._scroll_worker,
                        worker_id   = id_offset,
                        keyword     = keyword,
                        regiao      = regiao,
                        coords      = coords,
                        link_queue  = link_queue,
                        scroll_done = scroll_done,
                        results_lock= results_lock,
                        global_vistos= global_vistos,
                        headless    = headless,
                    )

                    # N ficha workers (W1..WN da keyword)
                    ficha_futures = [
                        ex.submit(
                            self._ficha_worker,
                            worker_id               = id_offset + 1 + fi,
                            regiao                  = regiao,
                            min_stars               = min_stars,
                            min_reviews             = min_reviews,
                            meta_por_kw             = meta_por_kw,
                            keyword                 = keyword,
                            link_queue              = link_queue,
                            scroll_done             = scroll_done,
                            results_lock            = results_lock,
                            global_vistos           = global_vistos,
                            global_aprovados_counter= global_aprovados_counter,
                            meta_total              = meta_total,
                            aprovados_kw_counter    = aprovados_kw_counter,
                            headless                = headless,
                        )
                        for fi in range(ficha_workers)
                    ]

                    # Coleta resultados dos ficha workers
                    for fut in as_completed(ficha_futures):
                        try:
                            r = fut.result()
                            with results_lock:
                                self.results.extend(r)
                        except Exception as e:
                            self.log(f"   ❌ ficha worker erro: {e}", "erro")

                    # Para o scroll worker se meta atingida
                    self.stop_flag = True  # sinaliza parada
                    try: scroll_future.result(timeout=10)
                    except Exception: pass
                    self.stop_flag = False  # reseta para próxima keyword

                kw_total = aprovados_kw_counter[0]
                status = "🎉 meta atingida" if kw_total >= meta_por_kw else f"📌 {kw_total}/{meta_por_kw}"
                self.log(f"── {status} para {keyword!r} ──", "ok" if kw_total >= meta_por_kw else "warn")

        except Exception as e:
            self.log(f"❌ erro geral: {e}", "erro")
        finally:
            try: self.db.close()
            except Exception: pass

        if self.results and save_path:
            self._exportar_excel(save_path)

        self.progress(100, f"concluído! {len(self.results)} aprovados")
        return self.results

    # ── Excel ─────────────────────────────────────────────────────────────────
    def _exportar_excel(self, path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Empresas"

        hdr_fill  = PatternFill("solid", start_color="1a237e")
        hdr_font  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        alt_fill  = PatternFill("solid", start_color="e8eaf6")
        wht_fill  = PatternFill("solid", start_color="FFFFFF")
        link_font = Font(color="1565C0", underline="single", name="Arial", size=10)
        norm_font = Font(name="Arial", size=10)
        center    = Alignment(horizontal="center", vertical="center")
        left      = Alignment(horizontal="left", vertical="center", wrap_text=False)
        borda     = Border(
            bottom=Side(style="thin", color="CCCCCC"),
            right =Side(style="thin", color="CCCCCC"),
        )

        colunas  = ["Nome","E-mail","Telefone","WhatsApp","Categoria",
                    "Endereço","Site","Estrelas","Avaliações","Keyword","URL Maps"]
        larguras = [35, 35, 18, 45, 20, 45, 40, 10, 12, 20, 50]

        for c, (col, larg) in enumerate(zip(colunas, larguras), 1):
            cell = ws.cell(row=1, column=c, value=col)
            cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = center
            ws.column_dimensions[get_column_letter(c)].width = larg
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"

        for ri, emp in enumerate(self.results, 2):
            fill = wht_fill if ri % 2 == 0 else alt_fill
            for ci, key in enumerate(colunas, 1):
                val  = emp.get(key, "")
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.fill = fill; cell.border = borda; cell.alignment = left
                if key in ("WhatsApp","Site","URL Maps") and str(val).startswith("http"):
                    cell.hyperlink = val; cell.font = link_font
                elif key == "E-mail" and "@" in str(val):
                    cell.hyperlink = f"mailto:{val}"; cell.font = link_font
                else:
                    cell.font = norm_font

        ws2 = wb.create_sheet("Resumo")
        n = len(self.results)
        for ri, (a, b, titulo) in enumerate([
            ("📊 RESUMO DA EXTRAÇÃO", "", True),
            ("", "", False),
            ("Total de empresas",  f"=COUNTA(Empresas!A2:A{n+1})", False),
            ("Com e-mail",         f'=COUNTIF(Empresas!F2:F{n+1},"?*")', False),
            ("Com WhatsApp",       f'=COUNTIF(Empresas!E2:E{n+1},"http*")', False),
            ("Média de estrelas",  f"=AVERAGE(Empresas!H2:H{n+1})", False),
            ("Data da extração",   datetime.now().strftime("%d/%m/%Y %H:%M"), False),
        ], 1):
            ws2.cell(row=ri, column=1, value=a).font = Font(
                bold=True, size=13 if titulo else 11, name="Arial",
                color="1a237e" if titulo else "000000"
            )
            if b:
                ws2.cell(row=ri, column=2, value=b).font = Font(name="Arial", color="1565C0")
        ws2.column_dimensions["A"].width = 24
        ws2.column_dimensions["B"].width = 28
        wb.save(path)
        self.log(f"💾 Planilha salva: {path}", "ok")


# ════════════════════════════════════════════════════════════════════════════
#  INTERFACE GRÁFICA
# ════════════════════════════════════════════════════════════════════════════
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("maps-scraper")
        self.geometry("1120x780")
        self.minsize(900, 660)
        self.configure(bg=COR_BG)
        self.scraper       = None
        self.thread        = None
        self.keywords_list = []
        self.save_path     = tk.StringVar(
            value=str(Path.home() / "Desktop" / "empresas_maps.xlsx")
        )
        self._build_ui()
        self._check_deps()

    def _build_ui(self):
        top = tk.Frame(self, bg=COR_BG)
        top.pack(fill="x", padx=20, pady=(14,4))
        tk.Label(top, text="maps-scraper",
                 bg=COR_BG, fg=COR_ACCENT, font=("Courier",18,"bold")).pack(side="left")
        tk.Label(top, text="  // extrator de leads via Google Maps",
                 bg=COR_BG, fg=COR_SUBTEXTO, font=("Courier",9)).pack(side="left", pady=(8,0))

        main = tk.Frame(self, bg=COR_BG)
        main.pack(fill="both", expand=True, padx=20, pady=6)
        main.columnconfigure(0, weight=1)
        main.columnconfigure(1, weight=2)
        main.rowconfigure(0, weight=1)

        # ── Painel esquerdo ──────────────────────────────────────────────────
        esq = tk.Frame(main, bg=COR_CARD,
                       highlightbackground=COR_BORDA, highlightthickness=1)
        esq.grid(row=0, column=0, sticky="nsew", padx=(0,8))

        self._sec(esq, "[ config ]", pad_top=14)

        self._sec(esq, "keywords")
        fkw = tk.Frame(esq, bg=COR_CARD)
        fkw.pack(fill="x", padx=14, pady=(0,4))
        self.ent_kw = self._entry(fkw)
        self.ent_kw.pack(side="left", fill="x", expand=True)
        self.ent_kw.bind("<Return>", lambda e: self._add_kw())
        self._btn(fkw, "+ Add", self._add_kw).pack(side="left", padx=(6,0))

        flb = tk.Frame(esq, bg=COR_INPUT,
                       highlightbackground=COR_BORDA, highlightthickness=1)
        flb.pack(fill="x", padx=14, pady=(0,4))
        self.lb = tk.Listbox(flb, bg=COR_INPUT, fg=COR_TEXTO, font=("Courier",10),
                             height=5, selectbackground=COR_ACCENT, selectforeground="#0d0b09",
                             borderwidth=0, highlightthickness=0)
        self.lb.pack(side="left", fill="both", expand=True, padx=4, pady=4)
        sb = tk.Scrollbar(flb, command=self.lb.yview, bg=COR_CARD)
        sb.pack(side="right", fill="y")
        self.lb.config(yscrollcommand=sb.set)
        self._btn(esq, "🗑  Remover selecionada", self._rem_kw,
                  cor=COR_DANGER).pack(anchor="w", padx=14, pady=(0,8))

        self._sec(esq, "região")
        self.ent_reg = self._entry(esq,
            placeholder="Ex: Ceará, Tokyo Japan, New York USA, Caçapava SP")
        self.ent_reg.pack(fill="x", padx=14, pady=(0,8))

        self._sec(esq, "filtros")
        ff = tk.Frame(esq, bg=COR_CARD)
        ff.pack(fill="x", padx=14, pady=(0,8))
        ff.columnconfigure((0,1), weight=1)
        tk.Label(ff, text="min_stars  (0–5)", bg=COR_CARD,
                 fg=COR_SUBTEXTO, font=("Courier",8)).grid(row=0,column=0,sticky="w")
        tk.Label(ff, text="min_reviews", bg=COR_CARD,
                 fg=COR_SUBTEXTO, font=("Courier",8)).grid(row=0,column=1,sticky="w",padx=(8,0))
        self.sp_stars = self._spin(ff, 0, 5, 0.5, "0")
        self.sp_stars.grid(row=1, column=0, sticky="ew", pady=(2,0))
        self.sp_rev = self._spin(ff, 0, 999999, 10, "0")
        self.sp_rev.grid(row=1, column=1, sticky="ew", padx=(8,0), pady=(2,0))

        self._sec(esq, "meta / keyword")
        tk.Label(esq,
                 text="# vasculha quantos forem necessários",
                 bg=COR_CARD, fg=COR_SUBTEXTO, font=("Courier",8),
                 wraplength=220, justify="left").pack(anchor="w", padx=14)
        self.sp_max = self._spin(esq, 1, 9999, 5, "20")
        self.sp_max.pack(fill="x", padx=14, pady=(4,8))

        self._sec(esq, "workers  (parallelism)")
        fw = tk.Frame(esq, bg=COR_CARD)
        fw.pack(fill="x", padx=14, pady=(0,4))
        fw.columnconfigure((0,1), weight=1)
        tk.Label(fw, text="# ficha workers  (abre N+1 Chromes: 1 scroll + N fichas)",
                 bg=COR_CARD, fg=COR_SUBTEXTO, font=("Courier",8),
                 wraplength=210, justify="left").grid(row=0,column=0,columnspan=2,sticky="w")
        self.sp_workers = self._spin(fw, 1, 8, 1, "1")
        self.sp_workers.grid(row=1, column=0, sticky="ew", pady=(2,4))
        self.var_headless = tk.BooleanVar(value=False)
        tk.Checkbutton(fw, text="headless (sem janela)", variable=self.var_headless,
                       bg=COR_CARD, fg=COR_TEXTO, font=("Courier",8),
                       selectcolor=COR_INPUT, activebackground=COR_CARD,
                       relief="flat").grid(row=1, column=1, sticky="w", padx=(8,0))

        self._sec(esq, "output")
        fsv = tk.Frame(esq, bg=COR_CARD)
        fsv.pack(fill="x", padx=14, pady=(0,14))
        tk.Entry(fsv, textvariable=self.save_path, bg=COR_INPUT, fg=COR_SUBTEXTO,
                 insertbackground=COR_ACCENT, font=("Courier",9), relief="flat",
                 highlightbackground=COR_BORDA, highlightthickness=1
                 ).pack(side="left", fill="x", expand=True)
        self._btn(fsv, "...", self._pick_path).pack(side="left", padx=(4,0))

        # ── Painel direito ───────────────────────────────────────────────────
        dir_ = tk.Frame(main, bg=COR_BG)
        dir_.grid(row=0, column=1, sticky="nsew")
        dir_.rowconfigure(1, weight=1)
        dir_.columnconfigure(0, weight=1)

        fbt = tk.Frame(dir_, bg=COR_BG)
        fbt.grid(row=0, column=0, sticky="ew", pady=(0,8))
        self.btn_start = tk.Button(fbt, text=">>  RUN",
            command=self._iniciar, bg=COR_ACCENT, fg="white",
            font=("Courier",12,"bold"), activebackground="#a05e2a",
            relief="flat", cursor="hand2", pady=10, padx=20)
        self.btn_start.pack(side="left", fill="x", expand=True, padx=(0,6))
        self.btn_stop = tk.Button(fbt, text="//  STOP",
            command=self._parar, bg=COR_DANGER, fg=COR_TEXTO,
            font=("Courier",12,"bold"), activebackground="#8a2020",
            relief="flat", cursor="hand2", pady=10, padx=20, state="disabled")
        self.btn_stop.pack(side="left", padx=(0,6))
        tk.Button(fbt, text="cls", command=self._clear_log,
            bg=COR_CARD, fg=COR_SUBTEXTO, font=("Courier",10),
            relief="flat", cursor="hand2", pady=10, padx=12).pack(side="left")
        tk.Button(fbt, text="limpar historico", command=self._limpar_historico,
            bg="#884444", fg="white", font=("Courier",9),
            relief="flat", cursor="hand2", pady=10, padx=10).pack(side="right")

        fp = tk.Frame(dir_, bg=COR_BG)
        fp.grid(row=2, column=0, sticky="ew", pady=(6,0))
        self.lbl_prog = tk.Label(fp, text="idle",
                                 bg=COR_BG, fg=COR_SUBTEXTO, font=("Courier",9))
        self.lbl_prog.pack(anchor="w")
        self.prog_var = tk.DoubleVar()
        style = ttk.Style(); style.theme_use("clam")
        style.configure("G.Horizontal.TProgressbar",
            troughcolor=COR_INPUT, background=COR_ACCENT, bordercolor=COR_BORDA)
        ttk.Progressbar(fp, variable=self.prog_var, maximum=100,
            style="G.Horizontal.TProgressbar").pack(fill="x", pady=(4,0))

        flog = tk.Frame(dir_, bg=COR_LOG_BG,
                        highlightbackground=COR_BORDA, highlightthickness=1)
        flog.grid(row=1, column=0, sticky="nsew")
        tk.Label(flog, text="$ stdout", bg=COR_LOG_BG,
                 fg=COR_ACCENT, font=("Courier",10,"bold")).pack(anchor="w", padx=10, pady=(8,4))
        ftxt = tk.Frame(flog, bg=COR_LOG_BG)
        ftxt.pack(fill="both", expand=True, padx=6, pady=(0,6))
        self.txt = tk.Text(ftxt, bg=COR_LOG_BG, fg=COR_TEXTO,
                           font=("Courier",9), relief="flat",
                           wrap="word", state="disabled")
        self.txt.pack(side="left", fill="both", expand=True)
        sb2 = tk.Scrollbar(ftxt, command=self.txt.yview, bg=COR_CARD)
        sb2.pack(side="right", fill="y")
        self.txt.config(yscrollcommand=sb2.set)
        for tag, cor in [("info","#6699cc"),("ok","#88bb44"),("erro","#ee4444"),
                         ("warn","#ddaa33"),("sub","#888888"),("system","#aaaaaa")]:
            self.txt.tag_config(tag, foreground=cor)

    def _sec(self, p, txt, pad_top=6):
        f = tk.Frame(p, bg=COR_CARD)
        f.pack(fill="x", padx=14, pady=(pad_top,2))
        tk.Label(f, text=txt, bg=COR_CARD, fg=COR_TEXTO,
                 font=("Courier",9,"bold")).pack(side="left")

    def _entry(self, p, placeholder=""):
        e = tk.Entry(p, bg=COR_INPUT, fg=COR_TEXTO, insertbackground=COR_ACCENT,
                     font=("Courier",10), relief="flat",
                     highlightbackground=COR_BORDA, highlightthickness=1)
        if placeholder:
            e.insert(0, placeholder); e.config(fg=COR_SUBTEXTO)
            def fi(ev,w=e,ph=placeholder):
                if w.get()==ph: w.delete(0,"end"); w.config(fg=COR_TEXTO)
            def fo(ev,w=e,ph=placeholder):
                if not w.get(): w.insert(0,ph); w.config(fg=COR_SUBTEXTO)
            e.bind("<FocusIn>",fi); e.bind("<FocusOut>",fo)
        return e

    def _spin(self, p, from_, to, inc, val):
        s = tk.Spinbox(p, from_=from_, to=to, increment=inc,
                       bg=COR_INPUT, fg=COR_TEXTO, insertbackground=COR_ACCENT,
                       font=("Courier",10), buttonbackground=COR_BORDA, relief="flat",
                       highlightbackground=COR_BORDA, highlightthickness=1)
        s.delete(0,"end"); s.insert(0,val)
        return s

    def _btn(self, p, txt, cmd, cor=COR_ACCENT):
        return tk.Button(p, text=txt, command=cmd, bg=cor, fg="#0d0b09",
                         font=("Courier",9,"bold"), relief="flat", cursor="hand2",
                         padx=8, pady=4, activebackground=COR_BORDA, activeforeground=COR_TEXTO)

    def _add_kw(self):
        kw = self.ent_kw.get().strip()
        if kw and kw not in self.keywords_list:
            self.keywords_list.append(kw)
            self.lb.insert("end", f"  > {kw}")
            self.ent_kw.delete(0,"end")

    def _rem_kw(self):
        sel = self.lb.curselection()
        if sel:
            self.lb.delete(sel[0]); self.keywords_list.pop(sel[0])

    def _pick_path(self):
        p = filedialog.asksaveasfilename(defaultextension=".xlsx",
            filetypes=[("Excel","*.xlsx"),("Todos","*.*")],
            initialfile="empresas_maps.xlsx")
        if p: self.save_path.set(p)

    def _clear_log(self):
        self.txt.config(state="normal")
        self.txt.delete("1.0","end")
        self.txt.config(state="disabled")

    def _limpar_historico(self):
        db = HistoricoDB()
        total = db.total()
        db.close()
        if total == 0:
            messagebox.showinfo("Histórico", "O histórico já está vazio.")
            return
        resp = messagebox.askyesno(
            "Limpar histórico",
            f"Isso vai apagar {total} empresa(s) do histórico.\n"
            "Na próxima busca, elas poderão aparecer novamente.\n\nConfirmar?"
        )
        if resp:
            db = HistoricoDB()
            db.limpar()
            db.close()
            self.log(f"🗑 Histórico limpo ({total} registros removidos).", "warn")

    def log(self, msg, tag="info"):
        ts = datetime.now().strftime("%H:%M:%S")
        self.txt.config(state="normal")
        self.txt.insert("end", f"[{ts}] {msg}\n", tag)
        self.txt.see("end")
        self.txt.config(state="disabled")

    def set_progress(self, val, texto=""):
        self.prog_var.set(val)
        if texto: self.lbl_prog.config(text=texto)

    def _iniciar(self):
        if not self.keywords_list:
            messagebox.showwarning("Atenção","Adicione pelo menos uma palavra-chave!")
            return
        ph  = "Ex: Ceará, Tokyo Japan, New York USA, Caçapava SP"
        reg = self.ent_reg.get().strip()
        if not reg or reg == ph:
            messagebox.showwarning("Atenção","Digite a região para a busca!")
            return
        try:
            min_stars = float(self.sp_stars.get())
            min_revs  = int(self.sp_rev.get())
            meta      = int(self.sp_max.get())
            workers   = int(self.sp_workers.get())
        except ValueError:
            messagebox.showerror("Erro","Valores inválidos nos filtros."); return
        headless = self.var_headless.get()

        self.btn_start.config(state="disabled")
        self.btn_stop.config(state="normal")
        self.prog_var.set(0)
        self.lbl_prog.config(text="Iniciando Chrome...")

        self.log("═"*55, "system")
        self.log(f"🚀 {len(self.keywords_list)} keyword(s) | Região: {reg}", "system")
        self.log(f"   Filtros: ≥{min_stars}⭐ | ≥{min_revs} aval. | Meta: {meta}/kw | Workers: {workers}{'  headless' if headless else ''}", "system")
        self.log("═"*55, "system")

        self.scraper = MapsScraper(
            log_cb      = lambda m, t="info": self.after(0, self.log, m, t),
            progress_cb = lambda v, t: self.after(0, self.set_progress, v, t),
        )

        def run():
            self.scraper.scrape(
                keywords    = self.keywords_list[:],
                regiao      = reg,
                min_stars   = min_stars,
                min_reviews = min_revs,
                meta_por_kw = meta,
                save_path   = self.save_path.get(),
                num_workers = workers,
                headless    = headless,
            )
            self.after(0, self._done)

        self.thread = threading.Thread(target=run, daemon=True)
        self.thread.start()

    def _parar(self):
        if self.scraper:
            self.scraper.stop()
            self.log("⏹ Parando após empresa atual...", "warn")
            self.btn_stop.config(state="disabled")

    def _done(self):
        self.btn_start.config(state="normal")
        self.btn_stop.config(state="disabled")
        n = len(self.scraper.results) if self.scraper else 0
        self.log("═"*55, "system")
        self.log(f"🎉 Concluído! {n} empresas aprovadas na planilha.", "ok")
        if n > 0:
            self.log(f"📁 {self.save_path.get()}", "ok")
            messagebox.showinfo("Pronto!",
                f"✅ {n} empresas aprovadas!\n\n📁 Salvo em:\n{self.save_path.get()}")

    def _check_deps(self):
        if MISSING:
            self.log("⚠  DEPENDÊNCIAS FALTANDO:", "erro")
            self.log(f"   pip install {' '.join(MISSING)}", "warn")
        else:
            self.log("✅ Dependências OK.", "ok")
            self.log("💡 O número que você define é a META de aprovados na planilha.", "sub")
            self.log("   O programa vasculha quantos estabelecimentos forem necessários.", "sub")
            try:
                db = HistoricoDB()
                total = db.total()
                db.close()
                self.log(f"📦 Histórico: {total} empresa(s) já capturadas (não serão repetidas).", "sub")
            except Exception:
                pass


if __name__ == "__main__":
    if MISSING:
        print(f"Instale: pip install {' '.join(MISSING)}")
    App().mainloop()