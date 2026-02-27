"""
╔══════════════════════════════════════════════════════════════════╗
║       GOOGLE MAPS SCRAPER v3 — Selenium Edition                 ║
║                                                                  ║
║  INSTALAÇÃO (apenas 1 comando):                                  ║
║  pip install selenium webdriver-manager openpyxl requests        ║
║              beautifulsoup4 lxml                                 ║
║                                                                  ║
║  REQUISITO: Google Chrome instalado no PC (gratuito)            ║
╚══════════════════════════════════════════════════════════════════╝
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import time
import re
import unicodedata
from datetime import datetime
from pathlib import Path

# ── Verificação de dependências ──────────────────────────────────────────────
MISSING = []
try:
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.chrome.options import Options
    from webdriver_manager.chrome import ChromeDriverManager
    from selenium.webdriver.chrome.service import Service
except ImportError:
    MISSING.append("selenium webdriver-manager")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    MISSING.append("openpyxl")

try:
    import requests
    from bs4 import BeautifulSoup
except ImportError:
    MISSING.append("requests beautifulsoup4 lxml")

# ── Cores ────────────────────────────────────────────────────────────────────
COR_BG       = "#0f0f1a"
COR_CARD     = "#1a1a2e"
COR_BORDA    = "#16213e"
COR_ACCENT   = "#4a9eff"
COR_ACCENT2  = "#00d4aa"
COR_DANGER   = "#ff4757"
COR_TEXTO    = "#e8e8f0"
COR_SUBTEXTO = "#8888aa"
COR_INPUT    = "#0d0d1f"
COR_LOG_BG   = "#080810"


# ════════════════════════════════════════════════════════════════════════════
#  SCRAPER  — v4: JavaScript DOM + regex no texto, sem CSS frágil
# ════════════════════════════════════════════════════════════════════════════
class MapsScraper:
    def __init__(self, log_cb=None, progress_cb=None):
        self.log       = log_cb      or print
        self.progress  = progress_cb or (lambda v, t: None)
        self.stop_flag = False
        self.results   = []
        self.driver    = None

    def stop(self):
        self.stop_flag = True

    # ── Normaliza texto para comparação ──────────────────────────────────────
    @staticmethod
    def _norm(txt):
        import unicodedata
        txt = txt.lower()
        txt = unicodedata.normalize("NFD", txt)
        return "".join(c for c in txt if unicodedata.category(c) != "Mn")

    # ── E-mail via site ───────────────────────────────────────────────────────
    def _email_do_site(self, url):
        if not url or not url.startswith("http"):
            return ""
        try:
            r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=8)
            soup = BeautifulSoup(r.text, "lxml")
            # Prioriza mailto: links
            for a in soup.find_all("a", href=True):
                if a["href"].startswith("mailto:"):
                    e = a["href"][7:].split("?")[0].strip()
                    if "@" in e and "." in e.split("@")[1]:
                        return e
            # Fallback: regex no texto
            for e in re.findall(r"[\w.+\-]+@[\w.\-]+\.[a-zA-Z]{2,}", soup.get_text(" ")):
                dom = e.split("@")[1].lower()
                if not any(x in dom for x in ["example","wix","wordpress","seusite","domain"]):
                    return e
        except Exception:
            pass
        return ""

    # ── WhatsApp ──────────────────────────────────────────────────────────────
    def _whatsapp(self, tel):
        """
        Gera link wa.me/ usando o número exatamente como veio do Maps.
        Não assume nem adiciona código de país.
        """
        if not tel:
            return ""
        nums = re.sub(r"\D", "", tel)
        if len(nums) >= 7:
            return f"https://wa.me/{nums}"
        return ""

    # ── Geocodificação ────────────────────────────────────────────────────────
    def _geocodificar(self, regiao):
        try:
            r = requests.get(
                "https://nominatim.openstreetmap.org/search",
                params={"q": regiao, "format": "json", "limit": 1},
                headers={"User-Agent": "MapsScraper/4.0"},
                timeout=8,
            )
            data = r.json()
            if data:
                lat = float(data[0]["lat"])
                lon = float(data[0]["lon"])
                self.log(f"   📍 {data[0].get('display_name','')[:70]}", "sub")
                return lat, lon
        except Exception as e:
            self.log(f"   ⚠ Geocodificação falhou: {e}", "warn")
        return None

    # ── Monta URL do Maps centrada na região ─────────────────────────────────
    def _url_busca(self, keyword, regiao, coords, zoom=13):
        kw = requests.utils.quote(keyword)
        if coords:
            lat, lon = coords
            return f"https://www.google.com/maps/search/{kw}/@{lat},{lon},{zoom}z"
        q = requests.utils.quote(f"{keyword}, {regiao}")
        return f"https://www.google.com/maps/search/{q}"

    # ── Verifica se endereço pertence à região ────────────────────────────────
    def _na_regiao(self, endereco, regiao):
        if not endereco:
            return True  # sem endereço = benefício da dúvida
        end_n = self._norm(endereco)
        for parte in [p.strip() for p in regiao.split(",") if p.strip()]:
            palavras = [w for w in self._norm(parte).split() if len(w) > 2]
            if palavras and not all(w in end_n for w in palavras):
                return False
        return True

    # ── Chrome ────────────────────────────────────────────────────────────────
    def _init_driver(self):
        opts = Options()
        opts.add_argument("--lang=pt-BR")
        opts.add_argument("--start-maximized")
        opts.add_argument("--disable-blink-features=AutomationControlled")
        opts.add_experimental_option("excludeSwitches", ["enable-automation"])
        opts.add_experimental_option("useAutomationExtension", False)
        opts.add_argument("--disable-notifications")
        driver = webdriver.Chrome(
            service=Service(ChromeDriverManager().install()), options=opts
        )
        driver.execute_script(
            "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"
        )
        return driver

    # ════════════════════════════════════════════════════════════════════════
    #  COLETA DE LINKS — via JavaScript puro, não depende de seletores CSS
    # ════════════════════════════════════════════════════════════════════════
    def _coletar_links(self):
        """
        Usa JavaScript para pegar TODOS os hrefs que contêm /maps/place/
        diretamente do DOM. Não depende de classes CSS que mudam.
        Retorna set de URLs limpas.
        """
        try:
            hrefs = self.driver.execute_script("""
                var links = document.querySelectorAll('a[href*="/maps/place/"]');
                var result = [];
                links.forEach(function(a) {
                    var h = a.href || "";
                    // Filtra só links de estabelecimento (não fotos, não reviews)
                    if (h.includes("/maps/place/") && !h.includes("/photos/") && !h.includes("/reviews/")) {
                        // Limpa parâmetros de rastreamento mas mantém o link base
                        var clean = h.split("?")[0];
                        result.push(clean);
                    }
                });
                // Remove duplicatas
                return [...new Set(result)];
            """)
            return set(hrefs or [])
        except Exception:
            return set()

    def _scroll_lista(self):
        """
        Scroll no painel esquerdo de resultados do Maps.
        Tenta o painel lateral primeiro, fallback para scroll da janela.
        """
        try:
            # O painel de resultados do Maps tem role="main" ou é a div scrollável à esquerda
            scrolled = self.driver.execute_script("""
                // Tenta encontrar o painel de resultados scrollável
                var seletores = [
                    'div[role="feed"]',
                    'div[aria-label*="resultado"]',
                    'div[aria-label*="result"]',
                    'div.m6QErb[aria-label]',
                    'div.m6QErb',
                ];
                for (var i = 0; i < seletores.length; i++) {
                    var el = document.querySelector(seletores[i]);
                    if (el && el.scrollHeight > el.clientHeight) {
                        el.scrollBy(0, 3000);
                        return "feed:" + seletores[i];
                    }
                }
                // Fallback: scroll na página
                window.scrollBy(0, 3000);
                return "window";
            """)
            return scrolled
        except Exception:
            return "err"

    def _fim_de_lista(self):
        """Detecta se o Maps exibiu mensagem de fim de resultados."""
        try:
            body = self.driver.find_element(By.TAG_NAME, "body").text.lower()
            fim_msgs = [
                "chegou ao final", "you've reached the end",
                "fim da lista", "no more results",
                "nenhum resultado", "no results found",
                "não foram encontrados", "we couldn't find"
            ]
            return any(m in body for m in fim_msgs)
        except Exception:
            return False

    def _aguardar_resultados(self, timeout=12):
        """Aguarda aparecer pelo menos 1 link de estabelecimento na página."""
        inicio = time.time()
        while time.time() - inicio < timeout:
            if self._coletar_links():
                return True
            time.sleep(0.8)
        return False

    # ════════════════════════════════════════════════════════════════════════
    #  EXTRAÇÃO DE DADOS DA FICHA
    #  Estratégia: texto da página + regex (robusto a mudanças de HTML)
    # ════════════════════════════════════════════════════════════════════════
    def _extrair_ficha(self):
        d = self.driver

        # Aguarda o nome aparecer (H1)
        try:
            WebDriverWait(d, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "h1"))
            )
        except Exception:
            pass
        time.sleep(1.5)

        # ── Pega o texto completo da ficha ───────────────────────────────────
        try:
            page_text = d.find_element(By.TAG_NAME, "body").text
        except Exception:
            page_text = ""

        # ── Nome (H1 — o mais confiável) ─────────────────────────────────────
        nome = ""
        try:
            for sel in ["h1", "h1.DUwDvf", "h1.fontHeadlineLarge"]:
                els = d.find_elements(By.CSS_SELECTOR, sel)
                for el in els:
                    t = el.text.strip()
                    if t and len(t) > 1:
                        nome = t
                        break
                if nome:
                    break
        except Exception:
            pass

        # ── Categoria ────────────────────────────────────────────────────────
        # O Maps exibe a categoria logo abaixo do nome, em botão ou span pequeno
        categoria = ""
        try:
            for sel in ["button.DkEaL", "button[jsaction*='category']",
                        "span.mgr77e", ".fontBodyMedium button"]:
                els = d.find_elements(By.CSS_SELECTOR, sel)
                for el in els:
                    t = el.text.strip()
                    if t and len(t) < 80:
                        categoria = t
                        break
                if categoria:
                    break
        except Exception:
            pass

        # ── Estrelas — aria-label é o mais confiável ─────────────────────────
        stars = 0.0
        try:
            # Busca qualquer span/div com aria-label contendo "estrela" ou "star"
            for sel in [
                '[aria-label*="estrela"]',
                '[aria-label*="star"]',
                '[aria-label*="rating"]',
                '[aria-label*="nota"]',
            ]:
                els = d.find_elements(By.CSS_SELECTOR, sel)
                for el in els:
                    lbl = el.get_attribute("aria-label") or ""
                    m = re.search(r'(\d[,.]\d|[1-5])', lbl)
                    if m:
                        v = float(m.group(1).replace(",", "."))
                        if 1.0 <= v <= 5.0:
                            stars = v
                            break
                if stars:
                    break
        except Exception:
            pass

        # Fallback: regex no texto da página — procura padrão "4,5 (1.234)"
        if not stars:
            m = re.search(r'\b([1-5][,.]\d)\s*\(', page_text)
            if m:
                try:
                    stars = float(m.group(1).replace(",", "."))
                except Exception:
                    pass

        # ── Avaliações ───────────────────────────────────────────────────────
        reviews = 0
        try:
            for sel in [
                '[aria-label*="avalia"]',
                '[aria-label*="review"]',
                '[aria-label*="opini"]',
            ]:
                els = d.find_elements(By.CSS_SELECTOR, sel)
                for el in els:
                    lbl = el.get_attribute("aria-label") or ""
                    # "1.234 avaliações" ou "1,234 reviews"
                    m = re.search(r'([\d.,]+)\s*(?:avalia|review|opini)', lbl, re.I)
                    if m:
                        v = int(re.sub(r"\D", "", m.group(1)))
                        if v > 0:
                            reviews = v
                            break
                if reviews:
                    break
        except Exception:
            pass

        # Fallback: regex no texto — "(1.234 avaliações)" ou "(1,234)"
        if not reviews:
            # Padrão "(número)" próximo à nota
            m = re.search(r'\b[1-5][,.]\d\s*\(([\d.,]+)\)', page_text)
            if not m:
                m = re.search(r'([\d.]+)\s+avalia', page_text, re.I)
            if m:
                try:
                    reviews = int(re.sub(r"\D", "", m.group(1)))
                except Exception:
                    pass

        # ── Endereço ─────────────────────────────────────────────────────────
        endereco = ""
        try:
            # data-item-id="address" é o mais confiável historicamente
            el = d.find_element(By.CSS_SELECTOR, '[data-item-id="address"]')
            endereco = (el.get_attribute("aria-label") or el.text).strip()
            endereco = re.sub(r"(?i)^endere[çc]o[:\s]*", "", endereco).strip()
        except Exception:
            pass
        if not endereco:
            # Fallback: regex no texto da página — procura padrão de endereço BR
            m = re.search(
                r'(?:Rua|Av(?:enida)?|Al(?:ameda)?|R\.|Estrada|Rod(?:ovia)?|Praça)[^\n]{5,80}',
                page_text, re.I
            )
            if m:
                endereco = m.group(0).strip()

        # ── Telefone ─────────────────────────────────────────────────────────
        telefone = ""
        try:
            el = d.find_element(By.CSS_SELECTOR, '[data-item-id^="phone"]')
            telefone = (el.get_attribute("aria-label") or el.text).strip()
            telefone = re.sub(r"(?i)^telefone[:\s]*", "", telefone).strip()
        except Exception:
            pass
        if not telefone:
            # Regex universal de telefone (BR e internacional)
            m = re.search(r'(?:\+\d{1,3}[\s\-]?)?\(?\d{2,3}\)?[\s\-]?\d{4,5}[\s\-]\d{4}', page_text)
            if m:
                telefone = m.group(0).strip()

        # ── Site ─────────────────────────────────────────────────────────────
        site = ""
        try:
            el = d.find_element(By.CSS_SELECTOR, '[data-item-id="authority"]')
            site = el.get_attribute("href") or ""
        except Exception:
            pass
        if not site:
            try:
                # Qualquer link externo na ficha que não seja do google
                els = d.find_elements(By.CSS_SELECTOR, "a[href^='http']")
                for el in els:
                    href = el.get_attribute("href") or ""
                    if href.startswith("http") and "google" not in href and "maps" not in href:
                        site = href
                        break
            except Exception:
                pass

        return dict(
            nome=nome, categoria=categoria, endereco=endereco,
            telefone=telefone, site=site, stars=stars, reviews=reviews
        )

    # ════════════════════════════════════════════════════════════════════════
    #  SCRAPE PRINCIPAL
    # ════════════════════════════════════════════════════════════════════════
    def scrape(self, keywords, regiao, min_stars, min_reviews, meta_por_kw, save_path):
        self.stop_flag = False
        self.results   = []
        ja_vistos      = set()   # links já processados (todas as keywords)

        if MISSING:
            self.log(f"❌ Faltam dependências: pip install {chr(39).join(MISSING)}", "erro")
            return []

        try:
            self.log("🌐 Iniciando Chrome...", "info")
            self.driver = self._init_driver()
        except Exception as e:
            self.log(f"❌ Erro ao abrir Chrome: {e}", "erro")
            return []

        # Geocodifica UMA vez
        self.log(f"🌎 Geocodificando: \"{regiao}\"...", "info")
        coords = self._geocodificar(regiao)
        if coords:
            self.log(f"   ✅ lat={coords[0]:.4f} lon={coords[1]:.4f}", "ok")
        else:
            self.log("   ⚠ Sem coordenadas — usando fallback textual", "warn")

        total_kw = len(keywords)

        try:
            for ki, keyword in enumerate(keywords):
                if self.stop_flag:
                    break

                self.log(f"\n🔍 [{ki+1}/{total_kw}] \"{keyword}\" em \"{regiao}\"", "info")
                self.log(f"   🎯 Meta: {meta_por_kw} aprovados | ≥{min_stars}⭐ | ≥{min_reviews} aval.", "info")

                aprovados_kw = 0
                processados  = set()   # links processados nesta keyword

                # Zooms em ordem crescente de abrangência
                # Para cidade/bairro começa em 14, para estado/país em 9
                if "," in regiao or len(regiao.split()) >= 3:
                    zooms = [14, 12, 10, 8]   # cidade → região → estado → país
                else:
                    zooms = [10, 8, 6]         # estado → país → continental

                for zoom in zooms:
                    if self.stop_flag or aprovados_kw >= meta_por_kw:
                        break

                    url = self._url_busca(keyword, regiao, coords, zoom=zoom)
                    zoom_label = {14:"bairro/cidade",12:"região",10:"estado",8:"país",6:"continental"}.get(zoom, str(zoom))
                    self.log(f"   🔭 Zoom: {zoom_label} | {url[:80]}", "sub")
                    self.driver.get(url)
                    time.sleep(3)

                    # Verifica se resultados carregaram
                    if not self._aguardar_resultados(timeout=12):
                        self.log(f"   ⚠ Nenhum resultado carregou (zoom {zoom}). Próximo zoom...", "warn")
                        continue

                    # ── Scroll + coleta de links ─────────────────────────────
                    links_da_pagina   = set()
                    sem_novos_scroll  = 0
                    MAX_SEM_NOVOS     = 10  # tentativas sem novos links antes de ir pro próximo zoom

                    while not self.stop_flag and aprovados_kw < meta_por_kw and sem_novos_scroll < MAX_SEM_NOVOS:

                        # Coleta links visíveis agora
                        novos_links = self._coletar_links() - links_da_pagina - processados - ja_vistos

                        if novos_links:
                            sem_novos_scroll = 0
                            links_da_pagina.update(novos_links)
                            self.log(f"   📥 +{len(novos_links)} links | total na página: {len(links_da_pagina)}", "sub")

                            # ── Processa cada novo link ──────────────────────
                            for link in list(novos_links):
                                if self.stop_flag or aprovados_kw >= meta_por_kw:
                                    break

                                processados.add(link)

                                try:
                                    self.driver.get(link)
                                    dados = self._extrair_ficha()

                                    nome    = dados["nome"] or "(sem nome)"
                                    stars   = dados["stars"]
                                    reviews = dados["reviews"]
                                    end     = dados["endereco"]

                                    # ── Filtro de localização ────────────────
                                    if not self._na_regiao(end, regiao):
                                        self.log(f"   🚫 {nome} — fora da região ({end[:50]})", "warn")
                                        # Volta à página de resultados
                                        self.driver.back()
                                        time.sleep(1.5)
                                        continue

                                    # ── Filtros de qualidade ─────────────────
                                    if stars > 0 and stars < min_stars:
                                        self.log(f"   ⏭ {nome} | {stars:.1f}⭐ < {min_stars} — pula", "sub")
                                        self.driver.back()
                                        time.sleep(1.5)
                                        continue
                                    if reviews > 0 and reviews < min_reviews:
                                        self.log(f"   ⏭ {nome} | {reviews} aval. < {min_reviews} — pula", "sub")
                                        self.driver.back()
                                        time.sleep(1.5)
                                        continue

                                    # ── APROVADO ─────────────────────────────
                                    ja_vistos.add(link)
                                    aprovados_kw += 1

                                    email = ""
                                    if dados["site"]:
                                        self.log(f"   📧 {dados['site'][:45]}...", "sub")
                                        email = self._email_do_site(dados["site"])

                                    self.results.append({
                                        "Nome":       nome,
                                        "Categoria":  dados["categoria"],
                                        "Endereço":   end,
                                        "Telefone":   dados["telefone"],
                                        "WhatsApp":   self._whatsapp(dados["telefone"]),
                                        "E-mail":     email,
                                        "Site":       dados["site"],
                                        "Estrelas":   stars,
                                        "Avaliações": reviews,
                                        "Keyword":    keyword,
                                        "URL Maps":   link,
                                    })

                                    meta_total    = total_kw * meta_por_kw
                                    aprovados_tot = ki * meta_por_kw + aprovados_kw
                                    self.progress(
                                        min(aprovados_tot / meta_total * 100, 99),
                                        f"{len(self.results)} aprovados ({aprovados_kw}/{meta_por_kw} nesta keyword)"
                                    )

                                    st = f"{stars:.1f}⭐" if stars > 0 else "sem nota"
                                    rv = f"{reviews:,} aval.".replace(",", ".") if reviews > 0 else "sem aval."
                                    self.log(
                                        f"   ✅ [{aprovados_kw}/{meta_por_kw}] {nome} | {st} | {rv}"
                                        + (f" | 📧" if email else ""),
                                        "ok"
                                    )

                                    # Volta para a lista de resultados para continuar coletando
                                    self.driver.back()
                                    time.sleep(2)

                                except Exception as e:
                                    self.log(f"   ⚠ Erro: {e}", "warn")
                                    try:
                                        self.driver.back()
                                        time.sleep(1)
                                    except Exception:
                                        pass
                                    continue

                        else:
                            sem_novos_scroll += 1

                        if aprovados_kw >= meta_por_kw:
                            break

                        # Verifica fim de lista
                        if self._fim_de_lista():
                            self.log(f"   📌 Fim da lista Maps (zoom {zoom_label}). Tentando zoom mais aberto...", "sub")
                            break

                        # Scroll para carregar mais resultados
                        metodo = self._scroll_lista()
                        time.sleep(2)

                # Resumo da keyword
                if aprovados_kw >= meta_por_kw:
                    self.log(f"   🎉 Meta atingida! {aprovados_kw}/{meta_por_kw} aprovados.", "ok")
                else:
                    self.log(
                        f"   📌 Esgotado: {aprovados_kw}/{meta_por_kw} aprovados encontrados "
                        f"(não existem mais resultados que atendam os filtros nesta região).",
                        "warn"
                    )

        finally:
            try:
                self.driver.quit()
            except Exception:
                pass

        if self.results and save_path:
            self._exportar_excel(save_path)

        self.progress(100, f"Concluído! {len(self.results)} empresas aprovadas")
        return self.results

    # ── Excel ─────────────────────────────────────────────────────────────────
    def _exportar_excel(self, path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Empresas"

        hdr_fill  = PatternFill("solid", start_color="1a237e")
        hdr_font  = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        alt_fill  = PatternFill("solid", start_color="e8eaf6")
        wht_fill  = PatternFill("solid", start_color="FFFFFF")
        link_font = Font(color="1565C0", underline="single", name="Arial", size=10)
        norm_font = Font(name="Arial", size=10)
        center    = Alignment(horizontal="center", vertical="center")
        left      = Alignment(horizontal="left", vertical="center", wrap_text=False)
        borda     = Border(
            bottom=Side(style="thin", color="CCCCCC"),
            right =Side(style="thin", color="CCCCCC"),
        )

        colunas  = ["Nome","Categoria","Endereço","Telefone","WhatsApp",
                    "E-mail","Site","Estrelas","Avaliações","Keyword","URL Maps"]
        larguras = [35, 20, 45, 18, 45, 35, 40, 10, 12, 20, 50]

        for c, (col, larg) in enumerate(zip(colunas, larguras), 1):
            cell = ws.cell(row=1, column=c, value=col)
            cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = center
            ws.column_dimensions[get_column_letter(c)].width = larg
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = "A2"

        for ri, emp in enumerate(self.results, 2):
            fill = wht_fill if ri % 2 == 0 else alt_fill
            for ci, key in enumerate(colunas, 1):
                val  = emp.get(key, "")
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.fill = fill; cell.border = borda; cell.alignment = left
                if key in ("WhatsApp","Site","URL Maps") and str(val).startswith("http"):
                    cell.hyperlink = val; cell.font = link_font
                elif key == "E-mail" and "@" in str(val):
                    cell.hyperlink = f"mailto:{val}"; cell.font = link_font
                else:
                    cell.font = norm_font

        ws2 = wb.create_sheet("Resumo")
        n = len(self.results)
        for ri, (a, b, titulo) in enumerate([
            ("📊 RESUMO DA EXTRAÇÃO", "", True),
            ("", "", False),
            ("Total de empresas",  f"=COUNTA(Empresas!A2:A{n+1})", False),
            ("Com e-mail",         f'=COUNTIF(Empresas!F2:F{n+1},"?*")', False),
            ("Com WhatsApp",       f'=COUNTIF(Empresas!E2:E{n+1},"http*")', False),
            ("Média de estrelas",  f"=AVERAGE(Empresas!H2:H{n+1})", False),
            ("Data da extração",   datetime.now().strftime("%d/%m/%Y %H:%M"), False),
        ], 1):
            ws2.cell(row=ri, column=1, value=a).font = Font(
                bold=True, size=13 if titulo else 11, name="Arial",
                color="1a237e" if titulo else "000000"
            )
            if b:
                ws2.cell(row=ri, column=2, value=b).font = Font(name="Arial", color="1565C0")
        ws2.column_dimensions["A"].width = 24
        ws2.column_dimensions["B"].width = 28
        wb.save(path)
        self.log(f"💾 Planilha salva: {path}", "ok")


# ════════════════════════════════════════════════════════════════════════════
#  INTERFACE GRÁFICA
# ════════════════════════════════════════════════════════════════════════════
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("🗺 Google Maps Scraper v3")
        self.geometry("1120x780")
        self.minsize(900, 660)
        self.configure(bg=COR_BG)
        self.scraper       = None
        self.thread        = None
        self.keywords_list = []
        self.save_path     = tk.StringVar(
            value=str(Path.home() / "Desktop" / "empresas_maps.xlsx")
        )
        self._build_ui()
        self._check_deps()

    def _build_ui(self):
        top = tk.Frame(self, bg=COR_BG)
        top.pack(fill="x", padx=20, pady=(14,4))
        tk.Label(top, text="🗺  Google Maps Scraper",
                 bg=COR_BG, fg=COR_ACCENT, font=("Arial",20,"bold")).pack(side="left")
        tk.Label(top, text="  Capture leads automaticamente — Selenium Edition",
                 bg=COR_BG, fg=COR_SUBTEXTO, font=("Arial",10)).pack(side="left", pady=(6,0))

        main = tk.Frame(self, bg=COR_BG)
        main.pack(fill="both", expand=True, padx=20, pady=6)
        main.columnconfigure(0, weight=1)
        main.columnconfigure(1, weight=2)
        main.rowconfigure(0, weight=1)

        # ── Painel esquerdo ──────────────────────────────────────────────────
        esq = tk.Frame(main, bg=COR_CARD,
                       highlightbackground=COR_BORDA, highlightthickness=1)
        esq.grid(row=0, column=0, sticky="nsew", padx=(0,8))

        self._sec(esq, "⚙  CONFIGURAÇÕES", pad_top=14)

        self._sec(esq, "🏷  Palavras-chave")
        fkw = tk.Frame(esq, bg=COR_CARD)
        fkw.pack(fill="x", padx=14, pady=(0,4))
        self.ent_kw = self._entry(fkw)
        self.ent_kw.pack(side="left", fill="x", expand=True)
        self.ent_kw.bind("<Return>", lambda e: self._add_kw())
        self._btn(fkw, "+ Add", self._add_kw).pack(side="left", padx=(6,0))

        flb = tk.Frame(esq, bg=COR_INPUT,
                       highlightbackground=COR_BORDA, highlightthickness=1)
        flb.pack(fill="x", padx=14, pady=(0,4))
        self.lb = tk.Listbox(flb, bg=COR_INPUT, fg=COR_TEXTO, font=("Arial",10),
                             height=5, selectbackground=COR_ACCENT, selectforeground="white",
                             borderwidth=0, highlightthickness=0)
        self.lb.pack(side="left", fill="both", expand=True, padx=4, pady=4)
        sb = tk.Scrollbar(flb, command=self.lb.yview, bg=COR_CARD)
        sb.pack(side="right", fill="y")
        self.lb.config(yscrollcommand=sb.set)
        self._btn(esq, "🗑  Remover selecionada", self._rem_kw,
                  cor=COR_DANGER).pack(anchor="w", padx=14, pady=(0,8))

        self._sec(esq, "📍  Região / Cidade / País")
        self.ent_reg = self._entry(esq,
            placeholder="Ex: Ceará, Tokyo Japan, New York USA, Caçapava SP")
        self.ent_reg.pack(fill="x", padx=14, pady=(0,8))

        self._sec(esq, "⭐  Filtros de qualidade")
        ff = tk.Frame(esq, bg=COR_CARD)
        ff.pack(fill="x", padx=14, pady=(0,8))
        ff.columnconfigure((0,1), weight=1)
        tk.Label(ff, text="Mín. Estrelas (0–5)", bg=COR_CARD,
                 fg=COR_SUBTEXTO, font=("Arial",9)).grid(row=0,column=0,sticky="w")
        tk.Label(ff, text="Mín. Avaliações", bg=COR_CARD,
                 fg=COR_SUBTEXTO, font=("Arial",9)).grid(row=0,column=1,sticky="w",padx=(8,0))
        self.sp_stars = self._spin(ff, 0, 5, 0.5, "0")
        self.sp_stars.grid(row=1, column=0, sticky="ew", pady=(2,0))
        self.sp_rev = self._spin(ff, 0, 999999, 10, "0")
        self.sp_rev.grid(row=1, column=1, sticky="ew", padx=(8,0), pady=(2,0))

        self._sec(esq, "🎯  Meta de aprovados por keyword")
        tk.Label(esq,
                 text="Vasculha quantos forem necessários até atingir este número.",
                 bg=COR_CARD, fg=COR_SUBTEXTO, font=("Arial",8),
                 wraplength=220, justify="left").pack(anchor="w", padx=14)
        self.sp_max = self._spin(esq, 1, 9999, 5, "20")
        self.sp_max.pack(fill="x", padx=14, pady=(4,8))

        self._sec(esq, "💾  Salvar planilha em")
        fsv = tk.Frame(esq, bg=COR_CARD)
        fsv.pack(fill="x", padx=14, pady=(0,14))
        tk.Entry(fsv, textvariable=self.save_path, bg=COR_INPUT, fg=COR_TEXTO,
                 insertbackground=COR_TEXTO, font=("Arial",9), relief="flat",
                 highlightbackground=COR_BORDA, highlightthickness=1
                 ).pack(side="left", fill="x", expand=True)
        self._btn(fsv, "...", self._pick_path).pack(side="left", padx=(4,0))

        # ── Painel direito ───────────────────────────────────────────────────
        dir_ = tk.Frame(main, bg=COR_BG)
        dir_.grid(row=0, column=1, sticky="nsew")
        dir_.rowconfigure(1, weight=1)
        dir_.columnconfigure(0, weight=1)

        fbt = tk.Frame(dir_, bg=COR_BG)
        fbt.grid(row=0, column=0, sticky="ew", pady=(0,8))
        self.btn_start = tk.Button(fbt, text="▶  INICIAR SCRAPING",
            command=self._iniciar, bg=COR_ACCENT2, fg="#0f0f1a",
            font=("Arial",12,"bold"), activebackground="#00a882",
            relief="flat", cursor="hand2", pady=10, padx=20)
        self.btn_start.pack(side="left", fill="x", expand=True, padx=(0,6))
        self.btn_stop = tk.Button(fbt, text="⏹  PARAR",
            command=self._parar, bg=COR_DANGER, fg="white",
            font=("Arial",12,"bold"), activebackground="#cc2233",
            relief="flat", cursor="hand2", pady=10, padx=20, state="disabled")
        self.btn_stop.pack(side="left", padx=(0,6))
        tk.Button(fbt, text="🗑 Limpar", command=self._clear_log,
            bg=COR_CARD, fg=COR_SUBTEXTO, font=("Arial",10),
            relief="flat", cursor="hand2", pady=10, padx=12).pack(side="left")

        fp = tk.Frame(dir_, bg=COR_BG)
        fp.grid(row=2, column=0, sticky="ew", pady=(6,0))
        self.lbl_prog = tk.Label(fp, text="Aguardando...",
                                 bg=COR_BG, fg=COR_SUBTEXTO, font=("Arial",9))
        self.lbl_prog.pack(anchor="w")
        self.prog_var = tk.DoubleVar()
        style = ttk.Style(); style.theme_use("clam")
        style.configure("G.Horizontal.TProgressbar",
            troughcolor=COR_CARD, background=COR_ACCENT2, bordercolor=COR_BORDA)
        ttk.Progressbar(fp, variable=self.prog_var, maximum=100,
            style="G.Horizontal.TProgressbar").pack(fill="x", pady=(4,0))

        flog = tk.Frame(dir_, bg=COR_LOG_BG,
                        highlightbackground=COR_BORDA, highlightthickness=1)
        flog.grid(row=1, column=0, sticky="nsew")
        tk.Label(flog, text="📋  LOG DE EXECUÇÃO", bg=COR_LOG_BG,
                 fg=COR_ACCENT, font=("Arial",10,"bold")).pack(anchor="w", padx=10, pady=(8,4))
        ftxt = tk.Frame(flog, bg=COR_LOG_BG)
        ftxt.pack(fill="both", expand=True, padx=6, pady=(0,6))
        self.txt = tk.Text(ftxt, bg=COR_LOG_BG, fg=COR_TEXTO,
                           font=("Consolas",9), relief="flat",
                           wrap="word", state="disabled")
        self.txt.pack(side="left", fill="both", expand=True)
        sb2 = tk.Scrollbar(ftxt, command=self.txt.yview, bg=COR_CARD)
        sb2.pack(side="right", fill="y")
        self.txt.config(yscrollcommand=sb2.set)
        for tag, cor in [("info",COR_ACCENT),("ok",COR_ACCENT2),("erro",COR_DANGER),
                         ("warn","#ffa502"),("sub",COR_SUBTEXTO),("system","#c678dd")]:
            self.txt.tag_config(tag, foreground=cor)

    def _sec(self, p, txt, pad_top=6):
        f = tk.Frame(p, bg=COR_CARD)
        f.pack(fill="x", padx=14, pady=(pad_top,2))
        tk.Label(f, text=txt, bg=COR_CARD, fg=COR_TEXTO,
                 font=("Arial",9,"bold")).pack(side="left")

    def _entry(self, p, placeholder=""):
        e = tk.Entry(p, bg=COR_INPUT, fg=COR_TEXTO, insertbackground=COR_TEXTO,
                     font=("Arial",10), relief="flat",
                     highlightbackground=COR_BORDA, highlightthickness=1)
        if placeholder:
            e.insert(0, placeholder); e.config(fg=COR_SUBTEXTO)
            def fi(ev,w=e,ph=placeholder):
                if w.get()==ph: w.delete(0,"end"); w.config(fg=COR_TEXTO)
            def fo(ev,w=e,ph=placeholder):
                if not w.get(): w.insert(0,ph); w.config(fg=COR_SUBTEXTO)
            e.bind("<FocusIn>",fi); e.bind("<FocusOut>",fo)
        return e

    def _spin(self, p, from_, to, inc, val):
        s = tk.Spinbox(p, from_=from_, to=to, increment=inc,
                       bg=COR_INPUT, fg=COR_TEXTO, insertbackground=COR_TEXTO,
                       font=("Arial",10), buttonbackground=COR_CARD, relief="flat",
                       highlightbackground=COR_BORDA, highlightthickness=1)
        s.delete(0,"end"); s.insert(0,val)
        return s

    def _btn(self, p, txt, cmd, cor=COR_ACCENT):
        return tk.Button(p, text=txt, command=cmd, bg=cor, fg="white",
                         font=("Arial",9,"bold"), relief="flat", cursor="hand2",
                         padx=8, pady=4, activebackground="#2060bb", activeforeground="white")

    def _add_kw(self):
        kw = self.ent_kw.get().strip()
        if kw and kw not in self.keywords_list:
            self.keywords_list.append(kw)
            self.lb.insert("end", f"  🏷  {kw}")
            self.ent_kw.delete(0,"end")

    def _rem_kw(self):
        sel = self.lb.curselection()
        if sel:
            self.lb.delete(sel[0]); self.keywords_list.pop(sel[0])

    def _pick_path(self):
        p = filedialog.asksaveasfilename(defaultextension=".xlsx",
            filetypes=[("Excel","*.xlsx"),("Todos","*.*")],
            initialfile="empresas_maps.xlsx")
        if p: self.save_path.set(p)

    def _clear_log(self):
        self.txt.config(state="normal")
        self.txt.delete("1.0","end")
        self.txt.config(state="disabled")

    def log(self, msg, tag="info"):
        ts = datetime.now().strftime("%H:%M:%S")
        self.txt.config(state="normal")
        self.txt.insert("end", f"[{ts}] {msg}\n", tag)
        self.txt.see("end")
        self.txt.config(state="disabled")

    def set_progress(self, val, texto=""):
        self.prog_var.set(val)
        if texto: self.lbl_prog.config(text=texto)

    def _iniciar(self):
        if not self.keywords_list:
            messagebox.showwarning("Atenção","Adicione pelo menos uma palavra-chave!")
            return
        ph  = "Ex: Ceará, Tokyo Japan, New York USA, Caçapava SP"
        reg = self.ent_reg.get().strip()
        if not reg or reg == ph:
            messagebox.showwarning("Atenção","Digite a região para a busca!")
            return
        try:
            min_stars = float(self.sp_stars.get())
            min_revs  = int(self.sp_rev.get())
            meta      = int(self.sp_max.get())
        except ValueError:
            messagebox.showerror("Erro","Valores inválidos nos filtros."); return

        self.btn_start.config(state="disabled")
        self.btn_stop.config(state="normal")
        self.prog_var.set(0)
        self.lbl_prog.config(text="Iniciando Chrome...")

        self.log("═"*55, "system")
        self.log(f"🚀 {len(self.keywords_list)} keyword(s) | Região: {reg}", "system")
        self.log(f"   Filtros: ≥{min_stars}⭐ | ≥{min_revs} aval. | Meta: {meta}/keyword", "system")
        self.log("═"*55, "system")

        self.scraper = MapsScraper(
            log_cb      = lambda m, t="info": self.after(0, self.log, m, t),
            progress_cb = lambda v, t: self.after(0, self.set_progress, v, t),
        )

        def run():
            self.scraper.scrape(
                keywords    = self.keywords_list[:],
                regiao      = reg,
                min_stars   = min_stars,
                min_reviews = min_revs,
                meta_por_kw = meta,
                save_path   = self.save_path.get(),
            )
            self.after(0, self._done)

        self.thread = threading.Thread(target=run, daemon=True)
        self.thread.start()

    def _parar(self):
        if self.scraper:
            self.scraper.stop()
            self.log("⏹ Parando após empresa atual...", "warn")
            self.btn_stop.config(state="disabled")

    def _done(self):
        self.btn_start.config(state="normal")
        self.btn_stop.config(state="disabled")
        n = len(self.scraper.results) if self.scraper else 0
        self.log("═"*55, "system")
        self.log(f"🎉 Concluído! {n} empresas aprovadas na planilha.", "ok")
        if n > 0:
            self.log(f"📁 {self.save_path.get()}", "ok")
            messagebox.showinfo("Pronto!",
                f"✅ {n} empresas aprovadas!\n\n📁 Salvo em:\n{self.save_path.get()}")

    def _check_deps(self):
        if MISSING:
            self.log("⚠  DEPENDÊNCIAS FALTANDO:", "erro")
            self.log(f"   pip install {' '.join(MISSING)}", "warn")
        else:
            self.log("✅ Dependências OK.", "ok")
            self.log("💡 O número que você define é a META de aprovados na planilha.", "sub")
            self.log("   O programa vasculha quantos estabelecimentos forem necessários.", "sub")


if __name__ == "__main__":
    if MISSING:
        print(f"Instale: pip install {' '.join(MISSING)}")
    App().mainloop()